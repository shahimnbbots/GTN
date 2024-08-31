import math
import traceback
from datetime import datetime
import warnings
from tkinter import filedialog
import pandas as pd
from seleniumwire import webdriver as wire_webdriver
from selenium import webdriver as selenium_webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver import ChromeOptions as Options, ActionChains
from selenium.webdriver import ChromeService as Service
import time
import requests
from seleniumwire.thirdparty.mitmproxy.types import Path
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import tkinter as tk
import win32com.client as win32
import os
import shutil
import openpyxl
from PyPDF2 import PdfReader, PdfWriter
from sheetfu import SpreadsheetApp, Table

# from webdriver_manager.chrome import ChromeDriverManager

warnings.filterwarnings("ignore", message="wmf image format is not supported", category=UserWarning)


def add_data(data):
    print(data)
    sa = SpreadsheetApp('mailer-400406-83227f4a1b2d.json')
    spreadsheet = sa.open_by_id('1DUALWUvuJY89FPyjxzTv_6vj66wohceTChELc7Z7Zn8')
    sheet = spreadsheet.get_sheet_by_name('Sheet1')
    data_range = sheet.get_data_range()
    if "Date" in data and data["Date"] is not None:
        # Convert date to a string in the format YYYY-MM-DD
        data["Date"] = data["Date/Time"].strftime('%d-%m-%Y %H:%M:%S')

    # print(data)
    table = Table(data_range)
    table.add_one(data)
    table.commit()


def gtn(booking_number, workbook_cci, workbook_ctpat, workbook_ckj, format_value, country_value, ci_numbers):
    service = Service()
    options = Options()
    # options.add_argument("--headless")
    # options.add_argument("--no-sandbox")
    # options.add_argument("--window-size=1280,720")
    # options.add_argument("--disable-gpu")
    options.add_experimental_option('detach', True)
    # Use webdriver_manager to automatically download the correct version of ChromeDriver
    # service = Service(ChromeDriverManager().install())
    driver = wire_webdriver.Chrome(service=service, options=options)
    # driver = selenium_webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 30)
    try:
        driver.get("https://network.infornexus.com/login")
        username = driver.find_element(By.CSS_SELECTOR, 'input[name="userid"]')
        username.send_keys("balaji.pandu@shahi")
        password = driver.find_element(By.CSS_SELECTOR, 'input[name="uPassword"]')
        password.send_keys("SEPL@1234")
        login_button = driver.find_element(By.CSS_SELECTOR, 'input#loginButton')
        login_button.click()

        continue_button = wait.until(ec.presence_of_element_located((By.ID, 'submitbutton')))
        continue_button.click()
        time.sleep(5)

        applications_link = wait.until(ec.element_to_be_clickable((By.ID, 'navmenu__applications')))
        applications_link.click()

        invoice_optn = driver.find_element(By.XPATH, '//*[@id="navmenu__inprogressinvoices"]')
        invoice_optn.click()

        invoice_input_label = wait.until(ec.presence_of_element_located((By.XPATH, '//span[contains(text(), "Invoice Number")]/ancestor::label')))
        input_field = invoice_input_label.find_element(By.XPATH, '//*[@id="fixed-outer-filter-content"]/div/div/div[2]/div/div[2]/span/span[2]/span/input')

        booking_number_str = str(int(booking_number))
        if len(booking_number_str) == 11:
            booking_number_str += '0'
        input_field.send_keys(booking_number_str)
        print(booking_number_str)

        apply = driver.find_element(By.XPATH, '//*[@id="fixed-outer-filter-content"]/div/div/div[2]/span/button[2]')
        apply.click()
        time.sleep(3)

        table = driver.find_element(By.CSS_SELECTOR, '#active > div > span > span > div > div > div.flexresults.hoverFocuses.singleFocus > table')
        tbody = table.find_element(By.TAG_NAME, 'tbody')
        rows = tbody.find_elements(By.TAG_NAME, 'tr')

        # invoice_found = False
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, 'td')
            if len(cells) >= 4:
                invoice_issue_date = cells[1].text
                total_amount = cells[3].text
                print("Invoice Issue Date:", invoice_issue_date)
                print("Total Amount:", total_amount)
                print("--------------------")

                if invoice_issue_date:
                    try:
                        inv_number = wait.until(
                            ec.element_to_be_clickable((By.XPATH, f'//a[contains(text(), "{int(booking_number)}")]')))
                        action_chains = ActionChains(driver)
                        action_chains.double_click(inv_number).perform()
                    except NoSuchElementException:
                        print("Element not found")
                    try:
                        time.sleep(5)
                        # inv_date
                        inv_date_text = driver.find_element(By.XPATH,
                                                            '//*[@id="sqeIdDivContent"]/form/table[7]/tbody/tr/td[2]/table/tbody/tr[2]/td[2]')
                        inv_date = inv_date_text.text
                        parsed_date = datetime.strptime(inv_date, "%Y-%m-%d")
                        formatted_issue_date = parsed_date.strftime("%d-%b-%y")
                        print("Inv value:", formatted_issue_date)
                        # Net and Gross values
                        gross = driver.find_element(By.XPATH,
                                                    '//*[@id="sqeIdDivContent"]/form/table[21]/tbody/tr/td[2]/table/tbody/tr[2]/td[2]')
                        gross_value = gross.text.replace("KILOGRAMS", "KGS")
                        print(gross_value)
                        # net
                        net = driver.find_element(By.XPATH,
                                                  '//*[@id="sqeIdDivContent"]/form/table[21]/tbody/tr/td[2]/table/tbody/tr[2]/td[5]')
                        net_value = net.text.replace("KILOGRAMS", "KGS")
                        print(net_value)
                        # SA number
                        sa_number = driver.find_element(By.XPATH,
                                                        '//*[@id="sqeIdDivContent"]/form/table[7]/tbody/tr/td[2]/table/tbody/tr[9]/td[5]/font')
                        sa_value = sa_number.text
                        print(sa_value)
                        td_element_value = ""
                        # manufacture address
                        table = driver.find_element(By.XPATH,
                                                    '//*[@id="sqeIdDivContent"]/form/table[13]/tbody/tr/td[2]/table/tbody/tr[2]/td[1]/table')
                        tbody = table.find_element(By.TAG_NAME, 'tbody')
                        tr_elements = tbody.find_elements(By.TAG_NAME, 'tr')
                        for tr in tr_elements:
                            td_element = tr.find_element(By.TAG_NAME, 'td')
                            td_element_value += td_element.text.strip() + "\n"
                            print(td_element_value)
                        time.sleep(3)
                        # CCI
                        sheet_cci = workbook_cci['CCI']
                        for cell in ['H3', 'G60', 'D65', 'G73', 'H73']:
                            sheet_cci[cell].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                        if not sheet_cci['H3'].value:
                            sheet_cci['H3'] = formatted_issue_date
                        if not sheet_cci['D65'].value:
                            sheet_cci['D65'] = sa_value
                        if not sheet_cci['A74'].value:
                            sheet_cci['A74'] = booking_number_str
                        if not sheet_cci['A81'].value:
                            sheet_cci['A81'].alignment = Alignment(wrap_text=True)
                            sheet_cci['A81'] = td_element_value
                        if not sheet_cci['G73'].value:
                            sheet_cci['G73'] = net_value
                        if not sheet_cci['H73'].value:
                            sheet_cci['H73'] = gross_value
                        try:
                            # CTPAT
                            if format_value == 'PVH':
                                sheet_ctpat = workbook_ctpat['TOMMY format']
                                for cell in ['C21', 'D21', 'E6']:
                                    sheet_ctpat[cell].alignment = Alignment(horizontal='center', vertical='top',
                                                                            wrap_text=True)
                                if not sheet_ctpat['E6'].value:
                                    sheet_ctpat['E6'] = sa_value
                            elif format_value == 'CKJ':
                                # CKJ
                                sheet_ckj = workbook_ckj['CKJ FORMAT']
                                for cell in ['C16', 'D16', 'E7']:
                                    sheet_ckj[cell].alignment = Alignment(horizontal='center', vertical='top',
                                                                          wrap_text=True)
                                if not sheet_ckj['E7'].value:
                                    sheet_ckj['E7'] = sa_value
                        except Exception as e:
                            print("Error:", e)
                        time.sleep(8)
                        # to download the Commercial Invoice
                        hover_element = wait.until(ec.presence_of_element_located((By.XPATH,
                                                                                   '//*[@id="sqeIdDivContent"]/form/table[2]/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td[3]/table/tbody/tr/td[1]/div[2]')))
                        actions = ActionChains(driver)
                        actions.move_to_element(hover_element).perform()
                        time.sleep(5)

                        view_pdf_option = wait.until(ec.element_to_be_clickable(
                            (By.CSS_SELECTOR, "a[id*='_adjustHtmlDropDown'][href*='CommercialInvoice']")))
                        view_pdf_option.click()

                        time.sleep(5)

                        pdf_url = None
                        cookies = None

                        for request in driver.requests:
                            if 'CommercialInvoicePDF.jsp' in request.url and request.response:
                                pdf_url = request.url
                                print(f"URL: {request.url}")
                                print(f"Method: {request.method}")
                                print(f"Status Code: {request.response.status_code}")

                                cookies = {cookie['name']: cookie['value'] for cookie in driver.get_cookies()}
                                print("Captured Cookies:", cookies)
                                break

                        if pdf_url and cookies:
                            try:
                                headers = {
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36',
                                    'Accept': 'application/pdf',
                                    'Referer': 'https://network.infornexus.com/',
                                }

                                response = requests.get(pdf_url, headers=headers, cookies=cookies, stream=True)
                                response.raise_for_status()

                                final_output_path = f'D:/Renamed-PDF/{booking_number_str}-{ci_numbers[0]}-INV.pdf'
                                with open(final_output_path, 'wb') as f:
                                    f.write(response.content)

                                print('PDF downloaded and saved as:', final_output_path)
                                close_button = driver.find_element(By.CSS_SELECTOR, '#rsModalCloseButton0')
                                driver.execute_script("arguments[0].click();", close_button)
                            except requests.RequestException as e:
                                print('Error while downloading PDF:', e)
                                data = {
                                    "Booking Number": booking_number,
                                    "Error": f"Error while downloading PDF: INV",
                                    "Date/Time": datetime.now(),
                                }
                                add_data(data)
                                return False
                        else:
                            print("PDF URL or cookies not captured.")
                            data = {
                                "Booking Number": booking_number,
                                "Error": f"Error while downloading PDF: INV",
                                "Date/Time": datetime.now(),
                            }
                            add_data(data)
                            return False
                        time.sleep(5)
                        # to download the PL Invoice
                        applications_link = wait.until(ec.element_to_be_clickable((By.ID, 'navmenu__applications')))
                        applications_link.click()
                        pl_option = driver.find_element(By.XPATH, '//*[@id="navmenu__inprogresspackinglists"]')
                        pl_option.click()
                        invoice_input_label = wait.until(ec.presence_of_element_located(
                            (By.XPATH, '//span[contains(text(), "Packing List Ref Number")]/ancestor::label')))
                        input_field = invoice_input_label.find_element(By.XPATH,
                                                                       '//*[@id="fixed-outer-filter-content"]/div/div/div[2]/div/div[2]/span/span[2]/span/input')
                        input_field.send_keys(str(int(booking_number)))
                        apply = driver.find_element(By.XPATH,
                                                    '//*[@id="fixed-outer-filter-content"]/div/div/div[2]/span/button[2]')
                        apply.click()

                        try:
                            inv_number = wait.until(
                                ec.element_to_be_clickable(
                                    (By.XPATH, f'//a[contains(text(), "{int(booking_number)}")]')))
                            action_chains = ActionChains(driver)
                            action_chains.double_click(inv_number).perform()
                        except NoSuchElementException:
                            print("Element not found")
                        time.sleep(8)

                        hover_element = wait.until(
                            ec.presence_of_element_located((By.XPATH,
                                                            '//*[@id="sqeIdDivContent"]/form/table[2]/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td[4]/table/tbody/tr/td[1]/div[2]'))
                        )

                        actions = ActionChains(driver)
                        actions.move_to_element(hover_element).perform()

                        view_pdf_option = wait.until(
                            ec.element_to_be_clickable((By.XPATH, "//a[@href='javascript:loadPDF();']"))
                        )
                        view_pdf_option.click()
                        time.sleep(10)
                        # Capture the network traffic after clicking "View PDF"
                        pdf_url = None
                        cookies = None

                        for request in driver.requests:
                            if 'PackingManifestPDF.jsp' in request.url:
                                # Capture the URL
                                pdf_url = request.url

                                # Capture the response details
                                print(f"URL: {request.url}")
                                print(f"Method: {request.method}")
                                print(f"Status Code: {request.response.status_code}")
                                # Capture cookies
                                cookies = {cookie['name']: cookie['value'] for cookie in driver.get_cookies()}
                                print("Captured Cookies:", cookies)
                                break

                        # Optional: Re-download the PDF using the captured URL and cookies
                        if pdf_url and cookies:
                            try:
                                headers = {
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36',
                                    'Accept': 'application/pdf',
                                    'Referer': 'https://network.infornexus.com/',
                                }

                                # Perform the actual request to download the PDF
                                response = requests.get(pdf_url, headers=headers, cookies=cookies, stream=True)
                                response.raise_for_status()  # Check if the request was successful
                                # Save the PDF data to a new file in D:/
                                final_output_path = Path(f'D:/Renamed-PDF/{booking_number_str}-{ci_numbers[0]}-PL.pdf')
                                with open(final_output_path, 'wb') as f:
                                    f.write(response.content)
                                print('PDF downloaded and saved as:', final_output_path)
                                return True
                            except requests.RequestException as e:
                                print('Error while re-downloading PDF:', e)
                                data = {
                                    "Booking Number": booking_number,
                                    "Error": f"Error while downloading PDF: PL",
                                    "Date/Time": datetime.now(),
                                }
                                add_data(data)
                                return False
                        else:
                            print("PDF URL or cookies not captured.")
                            data = {
                                "Booking Number": booking_number,
                                "Error": f"Error while downloading PDF: PL",
                                "Date/Time": datetime.now(),
                            }
                            add_data(data)
                            return False
                    except Exception as e:
                        print("Error during invoice data extraction:", e)
                        return False
                else:
                    try:
                        inv_number = wait.until(
                            ec.element_to_be_clickable((By.XPATH, f'//a[contains(text(), "{int(booking_number)}")]')))
                        action_chains = ActionChains(driver)
                        action_chains.double_click(inv_number).perform()

                        time.sleep(3)

                        parent_element = wait.until(ec.presence_of_element_located(
                            (By.XPATH, '//td/a[contains(@href, "jumpToStep(\'AdditionalTerms\')")]')))
                        additional_terms_link = parent_element.find_element(By.XPATH,
                                                                            './span[contains(text(), "Additional Terms")]')
                        additional_terms_link.click()
                        print("Clicked on 'Additional Terms'.")

                        time.sleep(3)

                        try:
                            first_checkbox = driver.find_element(By.XPATH, "//tr[1]//input[@type='checkbox'][1]")
                            if not first_checkbox.is_selected():
                                first_checkbox.click()
                                print("Clicked the first checkbox.")
                            time.sleep(3)
                            third_checkbox = driver.find_element(By.XPATH, "//tr[3]//input[@type='checkbox'][1]")
                            if not third_checkbox.is_selected():
                                third_checkbox.click()
                                print("Clicked the third checkbox.")
                        except NoSuchElementException:
                            print("Could not find the specified checkboxes.")

                        time.sleep(3)

                        checkbox_table = wait.until(ec.presence_of_element_located((By.CLASS_NAME, 'pagecolor')))
                        rows = checkbox_table.find_elements(By.TAG_NAME, 'tr')

                        for row in rows:
                            try:
                                cell_text = row.find_element(By.TAG_NAME, 'font').text
                                if country_value == "USA":
                                    if any(text in cell_text for text in [
                                        'COMMERCIAL INVOICE COPY, STATING SHIPMENT AUTHORIZATION NUMBER, MANUFACTURER\'S NAME, MANUFACTURER\'S ADDRESS OR MID #, COUNTRY OF ORIGIN.',
                                        'PACKING LIST COPY.',
                                        'PACKING SUMMARY COPY.']):
                                        checkbox = row.find_element(By.XPATH, './/input[@type="checkbox"]')
                                        if not checkbox.is_selected():
                                            checkbox.click()
                                            print(f"Clicked checkbox for: {cell_text}")
                                elif country_value == "CAN":
                                    if any(text in cell_text for text in [
                                        'COMMERCIAL INVOICE COPY, STATING SHIPMENT AUTHORIZATION NUMBER, MANUFACTURER\'S NAME, MANUFACTURER\'S ADDRESS OR MID #, COUNTRY OF ORIGIN.',
                                        'CANADA CUSTOMS INVOICE (IF DESTINATION COUNTRY IS CANADA), STATING SHIPMENT AUTHORIZATION NUMBER, MANUFACTURER\'S NAME, MANUFACTURER\'S ADDRESS OR MID #, COUNTRY OF ORIGIN.',
                                        'PACKING LIST COPY.',
                                        'PACKING SUMMARY COPY.']):
                                        checkbox = row.find_element(By.XPATH, './/input[@type="checkbox"]')
                                        if not checkbox.is_selected():
                                            checkbox.click()
                                            print(f"Clicked checkbox for: {cell_text}")
                                            time.sleep(1)
                            except NoSuchElementException:
                                continue

                        try:
                            parent_element = wait.until(ec.presence_of_element_located(
                                (By.XPATH, '//td/a[contains(@href, "jumpToStep(\'Review\')")]')))
                            additional_terms_link = parent_element.find_element(By.XPATH,
                                                                                './span[contains(text(), "Preview")]')
                            additional_terms_link.click()
                            print("Clicked on 'Preview'.")
                        except NoSuchElementException:
                            print("Could not find 'Preview' link.")
                        time.sleep(5)
                        validate_button = driver.find_element(By.XPATH, "//input[@value='Validate']")
                        driver.execute_script("arguments[0].click();", validate_button)
                        time.sleep(10)
                        approve_button = driver.find_element(By.XPATH, '//*[@id="sqeIdDivContent"]/form/table[2]/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/input')
                        driver.execute_script("arguments[0].click();", approve_button)
                        time.sleep(5)
                        ok_button = driver.find_element(By.XPATH, '//*[@id="PasswordSignDoc"]/table/tbody/tr[3]/td/table/tbody/tr/td[3]/table/tbody/tr/td[1]/input')
                        driver.execute_script("arguments[0].click();", ok_button)
                        time.sleep(5)
                        driver.switch_to.alert.accept()
                        time.sleep(5)
                        #again go to invoice number search//////////////////
                        applications_link = wait.until(ec.element_to_be_clickable((By.ID, 'navmenu__applications')))
                        applications_link.click()

                        invoice_optn = driver.find_element(By.XPATH, '//*[@id="navmenu__inprogressinvoices"]')
                        invoice_optn.click()

                        invoice_input_label = wait.until(ec.presence_of_element_located(
                            (By.XPATH, '//span[contains(text(), "Invoice Number")]/ancestor::label')))
                        input_field = invoice_input_label.find_element(By.XPATH,
                                                                       '//*[@id="fixed-outer-filter-content"]/div/div/div[2]/div/div[2]/span/span[2]/span/input')

                        booking_number_str = str(int(booking_number))
                        if len(booking_number_str) == 11:
                            booking_number_str += '0'
                        input_field.send_keys(booking_number_str)
                        print(booking_number_str)

                        apply = driver.find_element(By.XPATH,
                                                    '//*[@id="fixed-outer-filter-content"]/div/div/div[2]/span/button[2]')
                        apply.click()
                        time.sleep(3)
                        try:
                            inv_number = wait.until(
                                ec.element_to_be_clickable(
                                    (By.XPATH, f'//a[contains(text(), "{int(booking_number)}")]')))
                            action_chains = ActionChains(driver)
                            action_chains.double_click(inv_number).perform()
                        except NoSuchElementException:
                            print("Element not found")
                        # inv_date
                        inv_date_text = driver.find_element(By.XPATH,
                                                            '//*[@id="sqeIdDivContent"]/form/table[7]/tbody/tr/td[2]/table/tbody/tr[2]/td[2]')
                        inv_date = inv_date_text.text
                        parsed_date = datetime.strptime(inv_date, "%Y-%m-%d")
                        formatted_issue_date = parsed_date.strftime("%d-%b-%y")
                        print("Inv value:", formatted_issue_date)
                        # Net and Gross values
                        gross = driver.find_element(By.XPATH,
                                                    '//*[@id="sqeIdDivContent"]/form/table[21]/tbody/tr/td[2]/table/tbody/tr[2]/td[2]')
                        gross_value = gross.text.replace("KILOGRAMS", "KGS")
                        print(gross_value)
                        # net
                        net = driver.find_element(By.XPATH,
                                                  '//*[@id="sqeIdDivContent"]/form/table[21]/tbody/tr/td[2]/table/tbody/tr[2]/td[5]')
                        net_value = net.text.replace("KILOGRAMS", "KGS")
                        print(net_value)
                        # SA number
                        sa_number = driver.find_element(By.XPATH,
                                                        '//*[@id="sqeIdDivContent"]/form/table[7]/tbody/tr/td[2]/table/tbody/tr[9]/td[5]/font')
                        sa_value = sa_number.text
                        print(sa_value)
                        td_element_value = ""
                        # manufacture address
                        table = driver.find_element(By.XPATH,
                                                    '//*[@id="sqeIdDivContent"]/form/table[13]/tbody/tr/td[2]/table/tbody/tr[2]/td[1]/table')
                        tbody = table.find_element(By.TAG_NAME, 'tbody')
                        tr_elements = tbody.find_elements(By.TAG_NAME, 'tr')
                        for tr in tr_elements:
                            td_element = tr.find_element(By.TAG_NAME, 'td')
                            td_element_value += td_element.text.strip() + "\n"
                            print(td_element_value)
                        time.sleep(3)
                        # CCI
                        sheet_cci = workbook_cci['CCI']
                        for cell in ['H3', 'G60', 'D65', 'G73', 'H73']:
                            sheet_cci[cell].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                        if not sheet_cci['H3'].value:
                            sheet_cci['H3'] = formatted_issue_date
                        if not sheet_cci['D65'].value:
                            sheet_cci['D65'] = sa_value
                        if not sheet_cci['A74'].value:
                            sheet_cci['A74'] = booking_number_str
                        if not sheet_cci['A81'].value:
                            sheet_cci['A81'].alignment = Alignment(wrap_text=True)
                            sheet_cci['A81'] = td_element_value
                        if not sheet_cci['G73'].value:
                            sheet_cci['G73'] = net_value
                        if not sheet_cci['H73'].value:
                            sheet_cci['H73'] = gross_value
                        try:
                            # CTPAT
                            if format_value == 'PVH':
                                sheet_ctpat = workbook_ctpat['TOMMY format']
                                for cell in ['C21', 'D21', 'E6']:
                                    sheet_ctpat[cell].alignment = Alignment(horizontal='center', vertical='top',
                                                                            wrap_text=True)
                                if not sheet_ctpat['E6'].value:
                                    sheet_ctpat['E6'] = sa_value
                            elif format_value == 'CKJ':
                                # CKJ
                                sheet_ckj = workbook_ckj['CKJ FORMAT']
                                for cell in ['C16', 'D16', 'E7']:
                                    sheet_ckj[cell].alignment = Alignment(horizontal='center', vertical='top',
                                                                          wrap_text=True)
                                if not sheet_ckj['E7'].value:
                                    sheet_ckj['E7'] = sa_value
                        except Exception as e:
                            print("Error:", e)
                        time.sleep(8)
                        # to download the Commercial Invoice
                        hover_element = wait.until(ec.presence_of_element_located((By.XPATH,
                                                                                   '//*[@id="sqeIdDivContent"]/form/table[2]/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td[3]/table/tbody/tr/td[1]/div[2]')))
                        actions = ActionChains(driver)
                        actions.move_to_element(hover_element).perform()
                        time.sleep(8)

                        view_pdf_option = wait.until(ec.element_to_be_clickable(
                            (By.CSS_SELECTOR, "a[id*='_adjustHtmlDropDown'][href*='CommercialInvoice']")))
                        view_pdf_option.click()

                        time.sleep(5)

                        pdf_url = None
                        cookies = None

                        for request in driver.requests:
                            if 'CommercialInvoicePDF.jsp' in request.url and request.response:
                                pdf_url = request.url
                                print(f"URL: {request.url}")
                                print(f"Method: {request.method}")
                                print(f"Status Code: {request.response.status_code}")

                                cookies = {cookie['name']: cookie['value'] for cookie in driver.get_cookies()}
                                print("Captured Cookies:", cookies)
                                break

                        if pdf_url and cookies:
                            try:
                                headers = {
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36',
                                    'Accept': 'application/pdf',
                                    'Referer': 'https://network.infornexus.com/',
                                }

                                response = requests.get(pdf_url, headers=headers, cookies=cookies, stream=True)
                                response.raise_for_status()

                                final_output_path = f'D:/Renamed-PDF/{booking_number_str}-{ci_numbers[0]}-INV.pdf'
                                with open(final_output_path, 'wb') as f:
                                    f.write(response.content)

                                print('PDF downloaded and saved as:', final_output_path)
                                close_button = driver.find_element(By.CSS_SELECTOR, '#rsModalCloseButton0')
                                driver.execute_script("arguments[0].click();", close_button)
                            except requests.RequestException as e:
                                print('Error while downloading PDF:', e)
                                data = {
                                    "Booking Number": booking_number,
                                    "Error": f"Error while downloading PDF: INV",
                                    "Date/Time": datetime.now(),
                                }
                                add_data(data)
                                return False
                        else:
                            print("PDF URL or cookies not captured.")
                            data = {
                                "Booking Number": booking_number,
                                "Error": f"Error while downloading PDF: INV",
                                "Date/Time": datetime.now(),
                            }
                            add_data(data)
                            return False
                        time.sleep(5)
                        # to download the PL Invoice
                        applications_link = wait.until(ec.element_to_be_clickable((By.ID, 'navmenu__applications')))
                        applications_link.click()
                        pl_option = driver.find_element(By.XPATH, '//*[@id="navmenu__inprogresspackinglists"]')
                        pl_option.click()
                        invoice_input_label = wait.until(ec.presence_of_element_located(
                            (By.XPATH, '//span[contains(text(), "Packing List Ref Number")]/ancestor::label')))
                        input_field = invoice_input_label.find_element(By.XPATH,
                                                                       '//*[@id="fixed-outer-filter-content"]/div/div/div[2]/div/div[2]/span/span[2]/span/input')
                        input_field.send_keys(str(int(booking_number)))
                        apply = driver.find_element(By.XPATH,
                                                    '//*[@id="fixed-outer-filter-content"]/div/div/div[2]/span/button[2]')
                        apply.click()

                        try:
                            inv_number = wait.until(
                                ec.element_to_be_clickable(
                                    (By.XPATH, f'//a[contains(text(), "{int(booking_number)}")]')))
                            action_chains = ActionChains(driver)
                            action_chains.double_click(inv_number).perform()
                        except NoSuchElementException:
                            print("Element not found")
                        time.sleep(8)

                        hover_element = wait.until(
                            ec.presence_of_element_located((By.XPATH,
                                                            '//*[@id="sqeIdDivContent"]/form/table[2]/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td[4]/table/tbody/tr/td[1]/div[2]'))
                        )

                        actions = ActionChains(driver)
                        actions.move_to_element(hover_element).perform()

                        view_pdf_option = wait.until(
                            ec.element_to_be_clickable((By.XPATH, "//a[@href='javascript:loadPDF();']"))
                        )
                        view_pdf_option.click()
                        time.sleep(10)
                        # Capture the network traffic after clicking "View PDF"
                        pdf_url = None
                        cookies = None

                        for request in driver.requests:
                            if 'PackingManifestPDF.jsp' in request.url:
                                # Capture the URL
                                pdf_url = request.url

                                # Capture the response details
                                print(f"URL: {request.url}")
                                print(f"Method: {request.method}")
                                print(f"Status Code: {request.response.status_code}")
                                # Capture cookies
                                cookies = {cookie['name']: cookie['value'] for cookie in driver.get_cookies()}
                                print("Captured Cookies:", cookies)
                                break

                        # Optional: Re-download the PDF using the captured URL and cookies
                        if pdf_url and cookies:
                            try:
                                headers = {
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36',
                                    'Accept': 'application/pdf',
                                    'Referer': 'https://network.infornexus.com/',
                                }

                                # Perform the actual request to download the PDF
                                response = requests.get(pdf_url, headers=headers, cookies=cookies, stream=True)
                                response.raise_for_status()  # Check if the request was successful

                                # Save the PDF data to a new file in D:/
                                final_output_path = Path(f'D:/Renamed-PDF/{booking_number_str}-{ci_numbers[0]}-PL.pdf')
                                with open(final_output_path, 'wb') as f:
                                    f.write(response.content)

                                print('PDF downloaded and saved as:', final_output_path)
                                return True

                            except requests.RequestException as e:
                                print('Error while downloading PDF:', e)
                                data = {
                                    "Booking Number": booking_number,
                                    "Error": f"Error while downloading PDF: PL",
                                    "Date/Time": datetime.now(),
                                }
                                add_data(data)
                                return False
                        else:
                            print("PDF URL or cookies not captured.")
                            data = {
                                "Booking Number": booking_number,
                                "Error": f"Error while downloading PDF: PL",
                                "Date/Time": datetime.now(),
                            }
                            add_data(data)
                            return False
                    except Exception as e:
                        print("Error during invoice creation:", e)
                        return False
            else:
                data = {
                    "Booking Number": booking_number,
                    "Error": f"Booking number not found",
                    "Date/Time": datetime.now(),
                }
                add_data(data)
                return False
    finally:
        driver.quit()


def scraping(user, passw, ci_number, workbook_cci, workbook_ctpat, workbook_ckj, file_path_ckj, file_path_cci,
             format_value, file_path_ctpat):
    service = Service()
    options = Options()
    # options.add_argument("--headless")
    # # options.add_argument("--no-sandbox")
    # # options.add_argument("--window-size=1280,720")
    # # options.add_argument("--disable-gpu")
    options.add_experimental_option('detach', True)
    driver = selenium_webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 10)
    driver.get("http://intranetn.shahi.co.in:8080/ShahiExportIntranet/login")
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    username.send_keys(user)
    password = driver.find_element(By.ID, "password")
    password.send_keys(passw)
    driver.find_element(By.ID, 'savebutton').click()
    time.sleep(2)
    handles = [driver.current_window_handle]
    driver.execute_script(
        "javascript:openMenuPage('null' , 'LG - Shahi Logistics Pre Shipment' , '10360' , 'F' , 'Applications'  );")
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    handles.append(driver.current_window_handle)
    driver.execute_script(
        f"javascript:openAccessPage('http://f1pla-production02.shahi.co.in:8080/ShahiLogisticsDiv/shahiwebpages/MvxExp/PRE/division.action?appName=LGPRE' , 'Logistic Logistics Pre Shipment    -  (Division)' , '12828' , 'R' , '{user}', 'N', '50015135', 'N'); ")
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    time.sleep(3)
    ci_number = str(ci_number)  # Convert to string
    spans = driver.find_elements(By.TAG_NAME, "span")
    for i in spans:
        if ci_number.startswith("10") and "100" in i.get_attribute("innerText"):
            i.click()
            break
        elif ci_number.startswith("37") and "370" in i.get_attribute("innerText"):
            i.click()
            break
        elif ci_number.startswith("34") and "340" in i.get_attribute("innerText"):
            i.click()
            break
    ul = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, "#tabs > ul")))
    a_list = ul.find_elements(By.TAG_NAME, "a")
    for i in a_list:
        if "Invoice Pre Shipment" in i.get_attribute("innerText"):
            driver.execute_script("arguments[0].click();", i)
            break

    frame = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, 'iframe[src *= "PREINVMVX.action"]')))
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    # Ci number
    ci_no = driver.find_element(By.XPATH, '//*[@id="searchcino"]')
    ci_no.send_keys(ci_number)

    # Go
    go_button = driver.find_element(By.ID, 'searchId')
    go_button.click()
    time.sleep(2)

    # Forwarder
    forwarder = driver.find_element(By.XPATH, '//*[@id="FWD_NAME"]')
    forwarder_value = forwarder.get_attribute("value").split('-')[0].strip()
    print(forwarder_value)
    # ETD date
    etd = driver.find_element(By.XPATH, '//*[@id="etd_date"]')
    etd_value = etd.get_attribute("value")
    parsed_date = datetime.strptime(etd_value, "%d/%m/%Y")
    formatted_etd_date = parsed_date.strftime("%d-%m-%Y")
    print(formatted_etd_date)
    # loading port
    loading = driver.find_element(By.XPATH, '//*[@id="LOADING_PORT"]')
    loading_value = loading.get_attribute("value")
    print(loading_value)
    img = driver.find_element(By.XPATH,
                              '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[3]/td[10]/a/img')
    driver.execute_script("arguments[0].click();", img)
    iframe = driver.find_element(By.ID, 'handlefrm')
    driver.switch_to.frame(iframe)
    input_field = driver.find_element(By.XPATH, '//*[@id="unitparam"]')
    input_field.send_keys(loading_value)
    img_2 = driver.find_element(By.XPATH, '/html/body/form/div/div[1]/div/table/tbody/tr/td[3]/img')
    img_2.click()
    name = driver.find_element(By.XPATH,
                               '/html/body/form/div/div[2]/div/table/tbody/tr/td/div/table/tbody/tr/td[2]')
    name_value = name.text.split('-')[0]
    print(name_value)
    close = driver.find_element(By.XPATH,
                                '/html/body/form/div/div[2]/div/table/tbody/tr/td/div/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", close)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    try:
        sheet_cci = workbook_cci['CCI']
        # Check if the cells are empty before filling them
        if not sheet_cci['D31'].value:
            # Header
            discharge = driver.find_element(By.XPATH, '//*[@id="DISCHARGE_DESC"]')
            to_value = discharge.get_attribute("value")
            print(to_value)
            sheet_cci['D31'] = to_value

        if not sheet_cci['E31'].value:
            discharge_country = driver.find_element(By.XPATH, '//*[@id="DIS_CNTRY_DESC"]')
            discharge_country_value = discharge_country.get_attribute("value")
            print(discharge_country_value)
            discharge_country_value_with_brackets = f"({discharge_country_value})"
            sheet_cci['E31'] = discharge_country_value_with_brackets

        if not sheet_cci['B31'].value:
            sheet_cci['B31'] = name_value
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
        #     picture2
        if not sheet_cci['A1'].value:
            img = Image('D:/PVH-USA/Picture2.png')
            sheet_cci.add_image(img, 'A1')
            img.anchor = 'A1'
            # picture3
        if not sheet_cci['H104'].value:
            img = Image('D:/PVH-USA/Picture3.png')
            sheet_cci.add_image(img, 'H104')
            img.anchor = 'H104'
        time.sleep(5)
        cartons = driver.find_element(By.XPATH, '//*[@id="CTNS"]')
        carton_value = cartons.get_attribute("value")
        print(carton_value)
        # COLINE/GST
        co_line = driver.find_element(By.XPATH, '//*[@id="ui-id-5"]/span')
        co_line.click()
        time.sleep(5)
        unique_texts = set()

        # Locate the table and extract data
        table = driver.find_element(By.XPATH, '//*[@id="tablea"]')
        driver.execute_script("arguments[0].scrollIntoView(true);", table)
        td_elements = table.find_elements(By.XPATH,
                                          '//tr[@bgcolor="#f2f9fb"]/td[@style="font-size:8pt;color:#9400D3;font-weight: bold;" and @align="right"][1]')

        if td_elements:
            total_quantity_element = td_elements[-1]
            total_quantity = total_quantity_element.text.strip()

            if not total_quantity.endswith('.0'):
                print("Total Quantity:", total_quantity)
            else:
                total_quantity = total_quantity.rstrip('0').rstrip('.')
                print("Total Quantity:", total_quantity)

            tbody = table.find_element(By.XPATH, '//*[@id="tablea"]/tbody[2]')
            tr_elements = tbody.find_elements(By.TAG_NAME, 'tr')
            print("Number of rows:", len(tr_elements))

            # Set to store unique texts
            existing_text = sheet_cci['B43'].value if sheet_cci['B43'].value else ""
            # Set to store unique descriptions
            unique_desc = set()

            # Iterate through table rows
            for tr in tr_elements:
                desc = tr.find_element(By.XPATH, '//*[@id="tab5"]//tbody/tr/td[8]').text.strip()
                # Check if the description is not already in unique_desc and not already in existing text
                if desc not in unique_desc and desc not in existing_text:
                    # Add the description to unique_desc
                    unique_desc.add(desc)

            new_text = existing_text + ("\n" if existing_text.strip() else "") + "\n".join(unique_desc)
            sheet_cci['B43'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            sheet_cci['B43'] = new_text

            tbody = table.find_element(By.XPATH, '//*[@id="tablea"]/tbody[2]')
            tr_elements = tbody.find_elements(By.TAG_NAME, 'tr')
            print("Number of rows:", len(tr_elements))
            unique_texts = set()
            for tr in tr_elements:
                buyer_po = tr.find_element(By.XPATH, '//*[@id="tab5"]//tbody/tr/td[5]').text.strip()
                buyer_style = tr.find_element(By.XPATH, '//*[@id="tab5"]//tbody/tr/td[6]').text.strip()
                price_fc = tr.find_element(By.XPATH, '//*[@id="tab5"]//tbody/tr/td[11]').text.strip()
                unique_texts.add((total_quantity, buyer_po, buyer_style, price_fc, carton_value))

            # Enable wrap text for the cells
            for cell in ['B43', 'H43', 'G43']:
                sheet_cci[cell].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            # Calculate sum of quantities from C10 to C15
            total_quantity = 0
            total_carton = 0
            for text in unique_texts:
                quantity = int(text[0])
                total_quantity += quantity
                # cartons count in CCI
                carton_val = int(text[4])
                total_carton += carton_val

            # Retrieve the existing value in cell C16 (if any) and add it to the total
            existing_value = sheet_cci['G60'].value or 0
            total_quantity += existing_value
            existing_value = sheet_cci['A43'].value or 0
            total_carton += existing_value
            # Put the total in cell C16
            sheet_cci['G60'] = total_quantity
            sheet_cci['A43'] = total_carton
            # Debugging: print the total quantity to verify
            print("Total Quantity:", total_quantity, "Total Carton:", total_carton)

            # Iterate over unique_texts and update sheet_cci
            for text in unique_texts:
                print("Text from td:", text)
                # Buyer PO
                next_row = 7
                while sheet_cci.cell(row=next_row, column=8).value:
                    next_row += 1
                existing_text = sheet_cci.cell(row=next_row, column=8).value if sheet_cci.cell(row=next_row,
                                                                                               column=8).value else ""
                new_text = existing_text + ("\n" if existing_text.strip() else "") + text[1].strip()
                sheet_cci.cell(row=next_row, column=8).value = new_text

                # Buyer po number
                next_row = 52
                while sheet_cci.cell(row=next_row, column=2).value:
                    next_row += 1
                existing_text = sheet_cci.cell(row=next_row, column=2).value if sheet_cci.cell(row=next_row,
                                                                                               column=2).value else ""
                new_text = existing_text + ("\n" if existing_text.strip() else "") + text[1].strip()
                sheet_cci.cell(row=next_row, column=2).value = new_text

                # Buyer Style
                next_row = 52
                while sheet_cci.cell(row=next_row, column=4).value:
                    next_row += 1
                existing_text = sheet_cci.cell(row=next_row, column=4).value if sheet_cci.cell(row=next_row,
                                                                                               column=4).value else ""
                new_text = existing_text + ("\n" if existing_text.strip() else "") + text[2].strip()
                sheet_cci.cell(row=next_row, column=4).value = new_text

                # price_FC
                next_row = 43
                while sheet_cci.cell(row=next_row, column=8).value:
                    next_row += 1
                existing_text = sheet_cci.cell(row=next_row, column=8).value if sheet_cci.cell(row=next_row,
                                                                                               column=8).value else ""
                new_text = existing_text + ("\n" if existing_text.strip() else "") + text[3].strip()
                sheet_cci.cell(row=next_row, column=8).value = new_text

                # total qty
                next_row = 43
                while sheet_cci.cell(row=next_row, column=7).value:
                    next_row += 1
                existing_text = sheet_cci.cell(row=next_row, column=7).value if sheet_cci.cell(row=next_row,
                                                                                               column=7).value else ""
                new_text = existing_text + ("\n" if existing_text.strip() else "") + text[0].strip()
                sheet_cci.cell(row=next_row, column=7).value = new_text

                # ci_numbers
                next_row = 76
                while sheet_cci.cell(row=next_row, column=1).value:
                    next_row += 1
                existing_text = sheet_cci.cell(row=next_row, column=1).value if sheet_cci.cell(
                    row=next_row,
                    column=1).value else ""
                new_text = existing_text + ("\n" if existing_text.strip() else "") + ci_number.strip()
                sheet_cci.cell(row=next_row, column=1).value = new_text

            # Save the workbook
            workbook_cci.save(file_path_cci)
        try:
            if format_value == 'PVH':
                sheet_ctpat = workbook_ctpat['TOMMY format']
                for cell in ['A9', 'B9', 'C9', 'D9', 'E31', 'E29', 'E33', 'B2', 'E43', 'D51', 'D9']:
                    sheet_ctpat[cell].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                if not sheet_ctpat['E31'].value:
                    sheet_ctpat['E31'] = forwarder_value
                if not sheet_ctpat['E33'].value:
                    sheet_ctpat['E33'] = name_value
                if not sheet_ctpat['E29'].value:
                    sheet_ctpat['E29'] = formatted_etd_date
                    # picture1
                if not sheet_ctpat['B2'].value:
                    img = Image('D:/PVH-USA/Picture1.png')
                    sheet_ctpat.add_image(img, 'B2')
                    img.anchor = 'B2'
                    # picture5
                if not sheet_ctpat['E43'].value:
                    img = Image('D:/PVH-USA/Picture5.png')
                    sheet_ctpat.add_image(img, 'E43')
                    img.anchor = 'E43'
                    # picture6
                if not sheet_ctpat['D51'].value:
                    img = Image('D:/PVH-USA/Picture6.png')
                    sheet_ctpat.add_image(img, 'D51')
                    img.anchor = 'D51'
                    # Calculate sum of quantities from C10 to C15
                    total_quantity = 0
                    total_carton = 0
                    for text in unique_texts:
                        quantity = int(text[0])
                        carton_val = int(text[4])
                        total_quantity += quantity
                        total_carton += carton_val
                    # Retrieve the existing value in cell C16 (if any) and add it to the total
                    existing_value = sheet_ctpat['C21'].value or 0
                    total_quantity += existing_value
                    existing_value = sheet_ctpat['D21'].value or 0
                    total_carton += existing_value
                    # Put the total in cell C16
                    sheet_ctpat['C21'] = total_quantity
                    sheet_ctpat['D21'] = total_carton
                    # Debugging: print the total quantity to verify
                    print("Total Quantity:", total_quantity)
                    print("Total Carton:", total_carton)
                # Iterate over unique_texts and update sheet_cci
                for text in unique_texts:
                    print("Text from td:", text)
                    # po_no
                    next_row_po = 9
                    while sheet_ctpat.cell(row=next_row_po, column=1).value:
                        next_row_po += 1
                    existing_text_po = sheet_ctpat.cell(row=next_row_po, column=1).value if sheet_ctpat.cell(
                        row=next_row_po,
                        column=1).value else ""
                    new_text_po = existing_text_po + ("\n" if existing_text_po.strip() else "") + text[1].strip()
                    sheet_ctpat.cell(row=next_row_po, column=1).value = new_text_po
                    # style
                    next_row_style = 9
                    while sheet_ctpat.cell(row=next_row_style, column=2).value:
                        next_row_style += 1
                    existing_text_style = sheet_ctpat.cell(row=next_row_style, column=2).value if sheet_ctpat.cell(
                        row=next_row_style,
                        column=2).value else ""
                    new_text_style = existing_text_style + ("\n" if existing_text_style.strip() else "") + text[
                        2].strip()
                    sheet_ctpat.cell(row=next_row_style, column=2).value = new_text_style
                    # qty
                    next_row_qty = 9
                    while sheet_ctpat.cell(row=next_row_qty, column=3).value:
                        next_row_qty += 1
                    existing_text_qty = sheet_ctpat.cell(row=next_row_qty, column=3).value if sheet_ctpat.cell(
                        row=next_row_qty,
                        column=3).value else ""
                    new_text_qty = existing_text_qty + ("\n" if existing_text_qty.strip() else "") + text[0].strip()
                    sheet_ctpat.cell(row=next_row_qty, column=3).value = new_text_qty
                    # carton_value
                    next_row = 9
                    while sheet_ctpat.cell(row=next_row, column=4).value:
                        next_row += 1
                    existing_text = sheet_ctpat.cell(row=next_row, column=4).value if sheet_ctpat.cell(row=next_row,
                                                                                                       column=4).value else ""
                    new_text = existing_text + ("\n" if existing_text.strip() else "") + text[4].strip()
                    sheet_ctpat.cell(row=next_row, column=4).value = new_text
                workbook_ctpat.save(file_path_ctpat)
            elif format_value == 'CKJ':
                # CKJ
                sheet_ckj = workbook_ckj['CKJ FORMAT']
                for cell in ['A10', 'B10', 'C10', 'D10', 'E26', 'E24', 'E28', 'D10']:
                    sheet_ckj[cell].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                if not sheet_ckj['E26'].value:
                    sheet_ckj['E26'] = forwarder_value
                if not sheet_ckj['E28'].value:
                    sheet_ckj['E28'] = name_value
                if not sheet_ckj['E24'].value:
                    sheet_ckj['E24'] = formatted_etd_date
                    # picture5
                if not sheet_ckj['E38'].value:
                    img = Image('D:/PVH-USA/Picture5.png')
                    sheet_ckj.add_image(img, 'E38')
                    img.anchor = 'E38'
                #     picture6
                if not sheet_ckj['D46'].value:
                    img = Image('D:/PVH-USA/Picture6.png')
                    sheet_ckj.add_image(img, 'D46')
                    img.anchor = 'D46'
                # Calculate sum of quantities from C10 to C15
                total_quantity = 0
                total_carton = 0
                for text in unique_texts:
                    quantity = int(text[0])
                    carton_val = int(text[4])
                    total_quantity += quantity
                    total_carton += carton_val
                # Retrieve the existing value in cell C16 (if any) and add it to the total
                existing_value = sheet_ckj['C16'].value or 0
                total_quantity += existing_value
                existing_value = sheet_ckj['D16'].value or 0
                total_carton += existing_value
                # Put the total in cell C16
                sheet_ckj['C16'] = total_quantity
                sheet_ckj['D16'] = total_carton
                # Debugging: print the total quantity to verify
                print("Total Quantity:", total_quantity)
                print("Total Carton:", total_carton)
                for text in unique_texts:
                    print("Text from td (PVH):", text)
                    # po_no
                    next_row = 10
                    while sheet_ckj.cell(row=next_row, column=1).value:
                        next_row += 1
                    existing_text = sheet_ckj.cell(row=next_row, column=1).value if sheet_ckj.cell(row=next_row,
                                                                                                   column=1).value else ""
                    new_text = existing_text + ("\n" if existing_text.strip() else "") + text[1].strip()
                    sheet_ckj.cell(row=next_row, column=1).value = new_text
                    # style
                    next_row = 10
                    while sheet_ckj.cell(row=next_row, column=2).value:
                        next_row += 1
                    existing_text = sheet_ckj.cell(row=next_row, column=2).value if sheet_ckj.cell(row=next_row,
                                                                                                   column=2).value else ""
                    new_text = existing_text + ("\n" if existing_text.strip() else "") + text[2].strip()
                    sheet_ckj.cell(row=next_row, column=2).value = new_text
                    # qty
                    next_row = 10
                    while sheet_ckj.cell(row=next_row, column=3).value:
                        next_row += 1
                    existing_text = sheet_ckj.cell(row=next_row, column=3).value if sheet_ckj.cell(row=next_row,
                                                                                                   column=3).value else ""
                    new_text = existing_text + ("\n" if existing_text.strip() else "") + text[0].strip()
                    sheet_ckj.cell(row=next_row, column=3).value = new_text
                    # carton_value
                    next_row = 10
                    while sheet_ckj.cell(row=next_row, column=4).value:
                        next_row += 1
                    existing_text = sheet_ckj.cell(row=next_row, column=4).value if sheet_ckj.cell(row=next_row,
                                                                                                   column=4).value else ""
                    new_text = existing_text + ("\n" if existing_text.strip() else "") + text[4].strip()
                    sheet_ckj.cell(row=next_row, column=4).value = new_text
                workbook_ckj.save(file_path_ckj)
        except Exception as e:
            print("Error:", e)
    except Exception as e:
        print("Error:", e)
    driver.quit()


def get_sort_key(filename):
    # Define priorities for each suffix
    priority_mapping = [
        ('-INV', 1),
        ('- CCI(USA)', 2),
        ('- CCI', 2),
        ('-PL', 3),
        ('- PVH(USA)', 4),
        ('- PVH', 4)
    ]
    # Set a default high value in case no match is found
    default_priority = 5
    # Check for each suffix in the filename
    for suffix, priority in priority_mapping:
        # Use string endswith to check for suffixes more accurately
        if filename.endswith(suffix + '.pdf'):
            print(f"File: {filename} - Matched Suffix: {suffix} - Priority: {priority}")
            return priority, filename

    print(f"File: {filename} - No Matching Suffix - Default Priority: {default_priority}")
    return default_priority, filename


# Function to merge PDFs
def merge_pdfs(pdf_list, output_path):
    pdf_writer = PdfWriter()
    for pdf in pdf_list:
        pdf_reader = PdfReader(pdf)
        for page_num in range(len(pdf_reader.pages)):
            pdf_writer.add_page(pdf_reader.pages[page_num])
    with open(output_path, 'wb') as out:
        pdf_writer.write(out)


# Choose the file using file dialog
filetypes = (('excel files', '*.xls'), ('excel files', '*.xlsx'), ('excel files', '*.ods'))
filename = filedialog.askopenfilename(
    title='Open excel',
    initialdir="/",
    filetypes=filetypes
)

# Read the data from the chosen file
df = pd.read_excel(filename)
ci_number_a = df['CI number'].tolist()
format_value_a = df['Category'].tolist()
booking_number_a = df['Booking Number'].tolist()
country_value_a = df['Country'].tolist()


def test_entry_bot():
    try:
        user = entry.get()
        passw = entry1.get()

        status_label = tk.Label(app, text="")
        status_label.grid(row=4, column=0, columnspan=4)
        app.update()
        opdict = {}
        current_booking_number = None
        current_format = None
        current_country = None
        file_path_ctpat = None
        file_path_ckj = None

        for i in range(len(ci_number_a)):
            ci_number = ci_number_a[i]
            booking_number = booking_number_a[i]
            format_value = format_value_a[i]
            country_value = country_value_a[i]

            if pd.notna(ci_number):
                if math.isnan(booking_number):  # Check for nan values
                    booking_number = current_booking_number
                else:
                    current_booking_number = booking_number
                    current_format = format_value
                    current_country = country_value
                if booking_number not in opdict:
                    opdict[booking_number] = {'CI_numbers': [ci_number], 'Category': current_format, 'Country': current_country}
                else:
                    opdict[booking_number]['CI_numbers'].append(ci_number)

        print(opdict)
        # Initialize Excel application
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False  # Set to True if you want Excel to be visible during execution

        for booking_number, data in opdict.items():
            ci_numbers = data['CI_numbers']
            format_value = data['Category']
            country_value = data['Country']
            # Delete files in the destination folder before processing the new booking number
            for root, dirs, files in os.walk('D:/Destination'):
                for file in files:
                    file_path = os.path.join(root, file)
                    os.remove(file_path)
                    time.sleep(5)
            # Open the 'CCI-NEW FORMAT (std.xlsx)' workbook
            file_name_cci = 'D:/PVH-USA/CCI-NEW FORMAT (std.xlsx).xlsx'
            destination_cci = 'D:/Destination'

            if os.path.exists(file_name_cci):
                shutil.copy(file_name_cci, destination_cci)
                time.sleep(3)
                file_path_cci = os.path.join(destination_cci, os.path.basename(file_name_cci))
                print("Source file copied successfully.")
                print("File path:", file_path_cci)

                if os.path.exists(file_path_cci):
                    print("Copied file exists:", file_path_cci)
                    workbook_cci = openpyxl.load_workbook(file_path_cci)
                    print("Worksheet names:", workbook_cci.sheetnames)

                    if format_value == 'PVH':
                        # Open the 'CTPAT FORMAT EXCEL - DIVE - PVH-KNIT.xlsx' workbook
                        source_file_ctpat = 'D:/PVH-USA/CTPAT FORMAT EXCEL - DIVE - PVH-KNIT.xlsx'
                        destination_ctpat = 'D:/Destination'

                        if os.path.exists(source_file_ctpat):
                            try:
                                shutil.copy(source_file_ctpat, destination_ctpat)
                                time.sleep(3)
                                file_path_ctpat = os.path.join(destination_ctpat, os.path.basename(source_file_ctpat))
                                print("Source file copied successfully.")
                                print("File path:", file_path_ctpat)

                                if os.path.exists(file_path_ctpat):
                                    print("Copied file exists:", file_path_ctpat)
                                    workbook_ctpat = openpyxl.load_workbook(file_path_ctpat)
                                    print("Worksheet names:", workbook_ctpat.sheetnames)
                                    workbook_ckj = openpyxl.Workbook()
                                    booking_number_int = int(booking_number)
                                    booking_number_str = str(booking_number_int)
                                    if len(booking_number_str) == 11:
                                        booking_number_str += '0'
                                    status_label.config(text=f"Processing Booking Number {booking_number_str}...")
                                    app.update()
                                    # gtn(booking_number, workbook_cci, workbook_ctpat, workbook_ckj, format_value, country_value)
                                    if gtn(booking_number, workbook_cci, workbook_ctpat, workbook_ckj, format_value, country_value, ci_numbers):
                                        time.sleep(2)
                                        for ci_number in ci_numbers:
                                            scraping(user, passw, ci_number, workbook_cci, workbook_ctpat, workbook_ckj,
                                                     file_path_ckj, file_path_cci, format_value, file_path_ctpat)
                                        renamed_folder = 'D:/Renamed-xlsx/'
                                        renamed_pdf_folder = 'D:/Renamed-PDF/'
                                        # Iterate over all files in the Destination folder (D:/Destination)
                                        for root, dirs, files in os.walk('D:/Destination'):
                                            for file in files:
                                                file_path = os.path.join(root, file)
                                                booking_number_int = int(booking_number)
                                                booking_number_str = str(booking_number_int)
                                                if len(booking_number_str) == 11:
                                                    booking_number_str += '0'
                                                format_part = file.split("-")[-2].strip().split(".")[0].replace('.xlsx', '')
                                                # first ci_number
                                                ci_number_first = ci_numbers[0]
                                                # Modify file naming based on country_value
                                                if country_value == 'USA':
                                                    new_file_name = f"{booking_number_str} - {ci_number_first} - {format_part}(USA).xlsx"
                                                    pdf_file_name = f"{booking_number_str} - {ci_number_first} - {format_part}(USA).pdf"
                                                else:
                                                    new_file_name = f"{booking_number_str} - {ci_number_first} - {format_part}.xlsx"
                                                    pdf_file_name = f"{booking_number_str} - {ci_number_first} - {format_part}.pdf"
                                                new_file_path = os.path.join(renamed_folder, new_file_name)
                                                shutil.copy(file_path, new_file_path)

                                                # Convert copied and renamed Excel files to PDF
                                                pdf_file_path = os.path.join(renamed_pdf_folder, pdf_file_name)
                                                # Use win32com for Excel to PDF conversion
                                                excel = win32.Dispatch("Excel.Application")
                                                wb = excel.Workbooks.Open(file_path)
                                                wb.ActiveSheet.ExportAsFixedFormat(0, pdf_file_path)
                                                wb.Close(SaveChanges=False)
                                                excel.Quit()
                                                # Move PDF file and remove Excel file
                                                new_pdf_file_path = os.path.join(renamed_pdf_folder, pdf_file_name)
                                                shutil.move(pdf_file_path, new_pdf_file_path.replace('.xlsx', '.pdf'))
                                                os.remove(file_path)

                                            print("Excel files converted to PDF and Excel files deleted successfully.")
                                        time.sleep(2)
                                        # Merge PDFs
                                        booking_dict = {}
                                        for file_name in os.listdir(renamed_pdf_folder):
                                            if file_name.endswith('.pdf'):
                                                booking_number = file_name.split('-')[0].strip()
                                                file_path = os.path.join(renamed_pdf_folder, file_name)

                                                if booking_number not in booking_dict:
                                                    booking_dict[booking_number] = []
                                                booking_dict[booking_number].append(file_path)

                                        for booking_number, pdf_files in booking_dict.items():
                                            if country_value == 'USA':
                                                # Exclude files with -CCI(USA) if country_value is USA
                                                filtered_files = [file for file in pdf_files if
                                                                  '- CCI(USA)' not in file]
                                                print("Filtered files (excluding -CCI(USA)):")
                                                for file in filtered_files:
                                                    print(file)
                                            else:
                                                # No filtering needed for other countries
                                                filtered_files = pdf_files

                                                # Sort files based on priority
                                            filtered_files.sort(key=get_sort_key)

                                            # Print the sorted file list for debugging
                                            print("Sorted PDF files order:")
                                            for pdf_file in filtered_files:
                                                print(pdf_file)

                                            # Define the path for the merged PDF
                                            ci_number_first = ci_numbers[0]
                                            merged_pdf_path = os.path.join('D:/Renamed-PDF',f'{booking_number}-{ci_number_first}-DOC.pdf')
                                            # Ensure that filtered_files is used for merging
                                            if filtered_files:  # Ensure there are files to merge
                                                merge_pdfs(filtered_files, merged_pdf_path)
                                                print(f'Merged PDF created: {merged_pdf_path}')
                                        time.sleep(2)
                                        # Create a folder with the booking number and move all related files into it
                                        try:
                                            # Create a folder with the name of the booking number
                                            booking_folder = os.path.join(renamed_pdf_folder, booking_number_str)
                                            os.makedirs(booking_folder, exist_ok=True)

                                            # Move all files related to the booking number into the folder
                                            for file_name in os.listdir(renamed_pdf_folder):
                                                file_path = os.path.join(renamed_pdf_folder, file_name)

                                                # Check if it's a file (not a directory) and contains the booking number
                                                if os.path.isfile(file_path) and booking_number_str in file_name:
                                                    new_file_path = os.path.join(booking_folder, file_name)
                                                    shutil.move(file_path, new_file_path)

                                            print(
                                                f"All files for booking number {booking_number_str} have been moved to folder: {booking_folder}")
                                            data = {
                                                "Booking Number": booking_number_str,
                                                "Status": f"Completed",
                                                "Date/Time": datetime.now(),
                                            }
                                            add_data(data)
                                        except Exception as e:
                                            print("Error while creating the booking folder or moving files:", e)
                                    else:
                                        print("gtn() did not return True or a valid condition.")
                            except Exception as e:
                                print("Error:", e)
                                data = {
                                    "Booking Number": booking_number,
                                    "Error": f"Error",
                                    "Date/Time": datetime.now(),
                                }
                                add_data(data)

                    elif format_value == 'CKJ':
                        # Open the 'CTPAT FORMAT EXCEL - DIV - CKJ-KNIT' workbook
                        source_file_ckj = 'D:/PVH-USA/CTPAT FORMAT EXCEL - DIV - CKJ-KNIT.xlsx'
                        destination_ckj = 'D:/Destination'

                        if os.path.exists(source_file_ckj):
                            try:
                                shutil.copy(source_file_ckj, destination_ckj)
                                time.sleep(3)
                                file_path_ckj = os.path.join(destination_ckj, os.path.basename(source_file_ckj))
                                print("Source file copied successfully.")
                                print("File path:", file_path_ckj)

                                if os.path.exists(file_path_ckj):
                                    print("Copied file exists:", file_path_ckj)
                                    workbook_ckj = openpyxl.load_workbook(file_path_ckj)
                                    print("Worksheet names:", workbook_ckj.sheetnames)
                                    workbook_ctpat = openpyxl.Workbook()
                                    booking_number_int = int(booking_number)
                                    booking_number_str = str(booking_number_int)
                                    if len(booking_number_str) == 11:
                                        booking_number_str += '0'
                                    status_label.config(text=f"Processing Booking Number {booking_number_str}...")
                                    app.update()

                                    # gtn(booking_number, workbook_cci, workbook_ctpat, workbook_ckj, format_value, country_value)
                                    if gtn(booking_number, workbook_cci, workbook_ctpat, workbook_ckj, format_value, country_value, ci_numbers):
                                        time.sleep(2)
                                        for ci_number in ci_numbers:
                                            scraping(user, passw, ci_number, workbook_cci, workbook_ctpat, workbook_ckj,
                                                     file_path_ckj, file_path_cci, format_value, file_path_ctpat)
                                        renamed_folder = 'D:/Renamed-xlsx/'
                                        renamed_pdf_folder = 'D:/Renamed-PDF/'
                                        # Iterate over all files in the Destination folder (D:/Destination)
                                        for root, dirs, files in os.walk('D:/Destination'):
                                            for file in files:
                                                file_path = os.path.join(root, file)
                                                booking_number_int = int(booking_number)
                                                booking_number_str = str(booking_number_int)
                                                if len(booking_number_str) == 11:
                                                    booking_number_str += '0'
                                                format_part = file.split("-")[-2].strip().split(".")[0].replace('.xlsx', '')
                                                # first ci_number
                                                ci_number_first = ci_numbers[0]
                                                # Modify file naming based on country_value
                                                if country_value == 'USA':
                                                    new_file_name = f"{booking_number_str} - {ci_number_first} - {format_part}(USA).xlsx"
                                                    pdf_file_name = f"{booking_number_str} - {ci_number_first} - {format_part}(USA).pdf"
                                                else:
                                                    new_file_name = f"{booking_number_str} - {ci_number_first} - {format_part}.xlsx"
                                                    pdf_file_name = f"{booking_number_str} - {ci_number_first} - {format_part}.pdf"
                                                new_file_path = os.path.join(renamed_folder, new_file_name)
                                                shutil.copy(file_path, new_file_path)

                                                # Convert copied and renamed Excel files to PDF
                                                pdf_file_path = os.path.join(renamed_pdf_folder, pdf_file_name)
                                                # Use win32com for Excel to PDF conversion
                                                excel = win32.Dispatch("Excel.Application")
                                                wb = excel.Workbooks.Open(file_path)
                                                wb.ActiveSheet.ExportAsFixedFormat(0, pdf_file_path)
                                                wb.Close(SaveChanges=False)
                                                excel.Quit()
                                                # Move PDF file and remove Excel file
                                                new_pdf_file_path = os.path.join(renamed_pdf_folder, pdf_file_name)
                                                shutil.move(pdf_file_path, new_pdf_file_path.replace('.xlsx', '.pdf'))
                                                os.remove(file_path)

                                            print("Excel files converted to PDF and Excel files deleted successfully.")
                                        time.sleep(2)
                                        # Merge PDFs
                                        booking_dict = {}
                                        for file_name in os.listdir(renamed_pdf_folder):
                                            if file_name.endswith('.pdf'):
                                                booking_number = file_name.split('-')[0].strip()
                                                file_path = os.path.join(renamed_pdf_folder, file_name)

                                                if booking_number not in booking_dict:
                                                    booking_dict[booking_number] = []
                                                booking_dict[booking_number].append(file_path)

                                        for booking_number, pdf_files in booking_dict.items():
                                            # Sort files based on priority
                                            pdf_files.sort(key=get_sort_key)
                                            print("Sorted PDF files order:")
                                            for pdf_file in pdf_files:
                                                print(pdf_file)

                                            # first ci_number
                                            ci_number_first = ci_numbers[0]
                                            merged_pdf_path = os.path.join(renamed_pdf_folder,f'{booking_number}-{ci_number_first}-DOC.pdf')
                                            merge_pdfs(pdf_files, merged_pdf_path)
                                            print(f'Merged PDF created: {merged_pdf_path}')
                                        time.sleep(2)
                                        # Create a folder with the booking number and move all related files into it
                                        try:
                                            # Create a folder with the name of the booking number
                                            booking_folder = os.path.join(renamed_pdf_folder, booking_number_str)
                                            os.makedirs(booking_folder, exist_ok=True)

                                            # Move all files related to the booking number into the folder
                                            for file_name in os.listdir(renamed_pdf_folder):
                                                file_path = os.path.join(renamed_pdf_folder, file_name)

                                                # Check if it's a file (not a directory) and contains the booking number
                                                if os.path.isfile(file_path) and booking_number_str in file_name:
                                                    new_file_path = os.path.join(booking_folder, file_name)
                                                    shutil.move(file_path, new_file_path)
                                            print(
                                                f"All files for booking number {booking_number_str} have been moved to folder: {booking_folder}")
                                            data = {
                                                "Booking Number": booking_number_str,
                                                "Status": f"Completed",
                                                "Date/Time": datetime.now(),
                                            }
                                            add_data(data)
                                        except Exception as e:
                                            print("Error while creating the booking folder or moving files:", e)
                                    else:
                                        print("gtn() did not return True or a valid condition.")
                            except Exception as e:
                                print("Error:", e)
                                data = {
                                    "Booking Number": booking_number,
                                    "Error": f"Error",
                                    "Date/Time": datetime.now(),
                                }
                                add_data(data)
        app.destroy()
    except Exception as e:
        with open('error_log.txt', 'a') as f:
            f.write(f"Error: {str(e)}\n")
            f.write(traceback.format_exc())


app = tk.Tk()
app.geometry("500x100")
app.title("PVH - USA DOCUMENT AUTOMATION")

label = tk.Label(app, text="Username")
label.grid(row=0, column=0)
entry = tk.Entry(app)
entry.grid(row=0, column=1)

label1 = tk.Label(app, text="Password")
label1.grid(row=0, column=2)
entry1 = tk.Entry(app)
entry1.grid(row=0, column=3)

button = tk.Button(app, text="Run Bot", command=test_entry_bot, width=10)
button.grid(row=3, column=1, columnspan=4)
app.mainloop()
